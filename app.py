import streamlit as st
from openpyxl import load_workbook
import openpyxl
import pandas as pd
from pathlib import Path
import streamlit_authenticator as stauth
import pickle

# Define the Streamlit app
def main():
    def app():
        authenticator.logout("Logout", "sidebar")
        st.header("Upload Files")

        # Add two file upload buttons
        file1 = st.file_uploader("Upload Customer Forecast File")
        file2 = st.file_uploader("Upload Open Order File")

        if file1 and file2 is not None:
            try:
                wb = openpyxl.load_workbook(file1, data_only=True)
                open_order = pd.read_excel(file2)

                df1 = clean_file1(wb)
                df2 = clean_file2(open_order)
                results = combine_files(df1, df2)

                results.set_index('SEC Code', inplace = True)
                # fill missing values with 0
                results.fillna(0, inplace=True)
                results = results.astype(int)
                results = results.applymap(lambda x: x if x >= 0 else 0)
                st.dataframe(results)
                csv = convert_df(results)

                st.download_button(
                    label="Download data as CSV",
                    data=csv,
                    file_name='large_df.csv',
                    mime='text/csv',
                )

            except ValueError as e:
                st.info("File Type Error")

    # --- USER AUTHENTICATION ---
    names = ["Infineon Admin", "Jiarrelyn"]
    usernames = ["infineon_admin", "jiarrelyn_99"]

    # Load hashed passwords
    file_path = Path(__file__).parent / "hashed_pw.pkl"
    with file_path.open("rb") as file:
        hashed_passwords = pickle.load(file)

    authenticator = stauth.Authenticate(names, usernames, hashed_passwords, "sales_dashboard", "abcdef")
    name, authentication_status, username = authenticator.login("Infineon Forecasting Web \n Login", "main")

    if authentication_status == False:
        st.error("Username/password is incorrect")

    if authentication_status == None:
        st.warning("Please enter your username and password")

    if authentication_status:
        page_names_to_funcs = {
            "Main Page": app
        }

        selected_page = st.sidebar.selectbox("Select a page", page_names_to_funcs.keys())
        page_names_to_funcs[selected_page]()
    
def clean_file1(wb):
    sheet = wb['F C S T _Updated']
    sheet = wb['F C S T _Updated']
    rows = sheet.iter_rows()

    # Find the row that contains "SEC Code"
    for row in rows:
        for cell in row:
            if cell.value == "SEC Code":
                start_row = row[0].row
                start_col = cell.column
                break

    # Select all the data to the right and bottom of the "SEC Code" cell
    data = sheet.iter_cols(min_row=start_row, min_col=start_col, max_row=sheet.max_row, max_col=sheet.max_column)
    rows_list = []

    # Loop through each row and get the values in the cells
    for row in data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)

    df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
    # df.transpose()
    df = df.set_index('SEC Code').T
    df.reset_index(inplace = True)
    df.columns.values[0] = 'SEC Code'
    select_columns = ['SEC Code', None, '전용성', 'Part Number', 'Category', 
              1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
              26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44,
              45, 46, 47, 48, 49, 50, 51, 52]
    
    columns_to_keep = [i for i in df.columns.tolist() if i in select_columns]
    get_min_week = min([i for i in columns_to_keep if isinstance(i, int)])
    col_name = [str(item) + 'W' if isinstance(item, int) else 'Category' if item is None else item for item in columns_to_keep]
    st.info(get_min_week)
    st.info(columns_to_keep)
    df = df[columns_to_keep]
    df.columns = col_name
    st.dataframe(df)
    
    first_row = df.loc[0, get_min_week:].values.tolist()
    date_str_list = [d.strftime('%d.%m.%Y') for d in first_row]
    df = df.loc[(df['Category'] == 'P O+ F C S T')]
    df.drop(columns = ['전용성', 'Part Number', 'Category'], inplace = True)
    df.columns.get_loc(get_min_week)
    df = df.set_index('SEC Code')
    df = df.loc[:, get_min_week:]
    df.columns = [date_str_list, df.columns.tolist()]
    df.reset_index(inplace = True)
    
    return df

def clean_file2(df):
    df = df[['Customer Material Number', 'Open Qty', 'Customer requested date']]
    df.columns = ['SEC Code', 'Quantity', 'Date']

    return df

def combine_files(df1, df2):
    import pandas as pd
    from datetime import datetime, timedelta

    # convert date strings to datetime objects
    df2['Date'] = pd.to_datetime(df2['Date'], format='%d/%m/%Y')

    # iterate over the columns of df1 except for the first one
    for col in df1.columns[1:]:
        
        # get the start and end dates for the current column
        start_date = datetime.strptime(col[0], '%d/%m/%Y')
        end_date = start_date + timedelta(days=7)
        
        # filter df2 for records between the start and end dates
        df2_filtered = df2[(df2['Date'] >= start_date) & (df2['Date'] < end_date)]
        
        # group the filtered df2 by SEC Code and sum the quantities
        df2_grouped = df2_filtered.groupby('SEC Code')['Quantity'].sum().reset_index()
        
        # iterate over the groups in df2_grouped
        for index, row in df2_grouped.iterrows():
            
            # get the SEC Code and Quantity for the current group
            sec_code = row['SEC Code']
            quantity = row['Quantity']
            # subtract the quantity from the corresponding cell in df1
            df1.loc[df1['SEC Code'] == sec_code, col] -= quantity
        
    return df1

@st.cache
def convert_df(df):
    # IMPORTANT: Cache the conversion to prevent computation on every rerun
    return df.to_csv().encode('utf-8')

# Run the app
if __name__ == "__main__":
    main()
